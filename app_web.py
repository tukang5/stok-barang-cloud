import io
import pandas as pd
import streamlit as st
from supabase import create_client, Client
import openpyxl
import openpyxl.styles
import openpyxl.utils

# Set konfigurasi halaman web
st.set_page_config(
    page_title="Sistem Manajemen Stok & Log Cloud", page_icon="📦", layout="wide"
)

# === 1. KONEKSI SUPABASE ===
def inisialisasi_supabase() -> Client:
    url: str = st.secrets["SUPABASE_URL"]
    key: str = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = inisialisasi_supabase()

# === 2. SISTEM KEAMANAN LISENSI & LOGIN DINAMIS ===
def sistem_login():
    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False
    if "user_role" not in st.session_state:
        st.session_state["user_role"] = "Owner"

    if not st.session_state["logged_in"]:
        # Cek apakah sudah ada akun pengguna terdaftar di database
        try:
            cek_user = supabase.table("pengguna").select("*").execute()
            ada_user = len(cek_user.data) > 0
        except Exception:
            ada_user = False

        col1, col2, col3 = st.columns(3)
        with col2:
            st.write("")
            
            # JIKA BELUM ADA AKUN (Aplikasi Baru Pertama Kali Dibuka Klien)
            if not ada_user:
                st.subheader("🔑 Aktivasi Kunci Lisensi Aplikasi Baru")
                st.info("Aplikasi belum diaktivasi. Silakan masukkan Kunci Lisensi resmi dari Developer.")
                
                input_lisensi = st.text_input("Masukkan Kunci Lisensi (License Key)")
                buat_user = st.text_input("Buat Username Baru untuk Toko Anda")
                buat_pass = st.text_input("Buat Password Baru", type="password")
                buat_role = st.selectbox("Pilih Hak Akses Peran (Role):", ["Owner", "Karyawan"])
                
                if st.button("Aktifkan Aplikasi ✨", type="primary", use_container_width=True):
                    if not input_lisensi or not buat_user or not buat_pass:
                        st.error("Semua kolom pengisian wajib diisi!")
                    else:
                        cek_lisensi = supabase.table("lisensi").select("*").ilike("kode_kunci", input_lisensi.strip()).execute()
                        
                        lisensi_valid = False
                        if cek_lisensi.data and len(cek_lisensi.data) > 0:
                         data_kunci = cek_lisensi.data[0]
                         if str(data_kunci.get("status", "")).lower() == "tersedia":
                             lisensi_valid = True
                        
                        if lisensi_valid:
                            try:
                                supabase.table("pengguna").insert({"username": buat_user.strip(), "password": buat_pass.strip(), "role": buat_role}).execute()
                                supabase.table("lisensi").update({"status": "Terpakai"}).ilike("kode_kunci", input_lisensi.strip()).execute()
                                
                                st.session_state["logged_in"] = True
                                st.session_state["user_role"] = buat_role
                                st.success("Aktivasi Sukses! Selamat Datang di Dashboard Toko Anda.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Gagal Registrasi: {str(e)}")
                        else:
                            st.error("Kunci Lisensi Salah atau sudah kadaluwarsa/terpakai!")
            # JIKA SUDAH ADA AKUN (Kondisi Normal)
            else:
                st.subheader("🔒 Silakan Login Terlebih Dahulu")
                username = st.text_input("Username Toko")
                password = st.text_input("Password", type="password")

                if st.button("Masuk 🔓", type="primary", use_container_width=True):
                    fitur_cek = supabase.table("pengguna").select("*").eq("username", username.strip()).eq("password", password.strip()).execute()
                    if fitur_cek.data and len(fitur_cek.data) > 0:
                        data_login = fitur_cek.data[0]
                        st.session_state["logged_in"] = True
                        st.session_state["user_role"] = data_login.get("role", "Owner")

                        st.success("Login Berhasil!")
                        st.rerun()
                    else:
                        st.error("Username atau Password salah!")
        return False
    return True

# === 3. FUNGSI LOGIKA DATABASE BARANG ===
def ambil_data(cari=""):
    try:
        if cari:
            response = supabase.table("barang").select("*").ilike("nama", f"%{cari}%").execute()
        else:
            response = supabase.table("barang").select("*").execute()
        
        df = pd.DataFrame(response.data)
        if df.empty:
            return pd.DataFrame(columns=["id", "nama", "stok", "harga_beli", "harga"])
        return df[["id", "nama", "stok", "harga_beli", "harga"]]
    except Exception:
        return pd.DataFrame(columns=["id", "nama", "stok", "harga_beli", "harga"])

def ambil_riwayat():
    try:
        response = supabase.table("riwayat").select("*").order("id", desc=True).execute()
        df = pd.DataFrame(response.data)
        if df.empty:
            return pd.DataFrame(columns=["waktu", "nama_barang", "tipe", "jumlah", "keterangan"])
        df["waktu"] = pd.to_datetime(df["waktu"]).dt.strftime('%Y-%m-%d %H:%M:%S')
        return df[["waktu", "nama_barang", "tipe", "jumlah", "keterangan"]]
    except Exception:
        return pd.DataFrame(columns=["waktu", "nama_barang", "tipe", "jumlah", "keterangan"])

def catat_log(nama_barang, tipe, jumlah, keterangan):
    try:
        supabase.table("riwayat").insert({
            "nama_barang": nama_barang, "tipe": tipe, "jumlah": jumlah, "keterangan": keterangan
        }).execute()
    except Exception:
        pass

# === 4. FUNGSI PENDUKUNG (FORMAT EXCEL OTOMATIS & REAL-TIME) ===
def konversi_ke_excel(df, sheet_name="Data"):
    output = io.BytesIO()
    
    # 1. Pastikan data yang ditarik selalu versi paling segar dari DataFrame Pandas
    df_Format = df.copy()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_Format.to_excel(writer, index=False, sheet_name=sheet_name, startrow=3)
        
        # Ambil kontrol workbook openpyxl untuk merapikan desain kolom
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # 2. Tambahkan Judul Laporan Formal di baris paling atas Excel
        worksheet["A1"] = f"LAPORAN RESMI PERSIDIAAN GUDANG ({sheet_name.upper()})"
        worksheet["A1"].font = openpyxl.styles.Font(name="Arial", size=14, bold=True, color="1A365D")
        worksheet["A2"] = "SISTEM MANAJEMEN STOK ENTERPRISE CLOUD REALT-TIME"
        worksheet["A2"].font = openpyxl.styles.Font(name="Arial", size=10, italic=True, color="4A5568")
        
        # 3. Desain Header Tabel (Warna Biru Navy Pro + Teks Putih Tebal)
        header_font = openpyxl.styles.Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        alignment_center = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # Warnai baris judul kolom (baris ke-4 di Excel karena startrow=3)
        for col_num in range(1, len(df_Format.columns) + 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
        
        # 4. Auto-Fit: Membuat lebar kolom otomatis melebar pas sesuai panjang teks (Anti-Terpotong/###)
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # 5. Berikan garis border tipis abu-abu agar tabel rapi saat dicetak di kertas
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='CBD5E0'),
            right=openpyxl.styles.Side(style='thin', color='CBD5E0'),
            top=openpyxl.styles.Side(style='thin', color='CBD5E0'),
            bottom=openpyxl.styles.Side(style='thin', color='CBD5E0')
        )
        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=len(df_Format.columns)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 4: # Mengatur isi data agar rata kiri teks, rata kanan angka
                    cell.alignment = openpyxl.styles.Alignment(vertical="center")

    return output.getvalue()
    
# === 5. LOGIKA UTAMA TAMPILAN APLIKASI ===
if sistem_login():
    st.sidebar.title("📌 Menu Navigasi")
    menu = st.sidebar.radio("Pilih Halaman:", ["Stok Barang Utama", "📋 Riwayat / Log Transaksi"])
    st.sidebar.write("---")
    st.sidebar.info(f"🎭 Hak Akses: **{st.session_state['user_role']}**")
    if st.sidebar.button("🚪 Keluar / Logout", use_container_width=True):
        st.session_state.clear()
        st.session_state["logged_in"] = False
        st.rerun()

    if menu == "Stok Barang Utama":
        st.title("📦 Sistem Manajemen Stok Barang (Enterprise Cloud)")
        st.markdown("Aplikasi manajemen stok berskala industri retail, grosir, dan gudang pabrik.")
        st.markdown("---")

        kolom_kiri, kolom_kanan = st.columns(2, gap="large")

        with kolom_kiri:
            st.subheader("📝 Formulir Barang")
            
            opsi_tindakan = ["Tambah Barang Baru", "Update Stok Masuk/Keluar"]
            if st.session_state["user_role"] == "Owner":
                opsi_tindakan.append("Hapus Barang")
                
            mode = st.radio("Pilih Tindakan:", opsi_tindakan)

            if mode == "Tambah Barang Baru":
                nama = st.text_input("Nama Barang")
                stok = st.number_input("Jumlah Stok Awal", min_value=0, step=1)
                harga_beli = st.number_input("Harga Beli / Modal Satuan (Rp)", min_value=0.0)
                harga = st.number_input("Harga Jual Satuan (Rp)", min_value=0.0)

                if st.button("➕ Tambah Barang", type="primary"):
                    if not nama:
                        st.error("Nama barang tidak boleh kosong!")
                    elif harga_beli > harga:
                        st.warning("⚠️ Peringatan: Harga beli lebih besar dari harga jual (Potensi Rugi)!")
                    else:
                        try:
                            supabase.table("barang").insert({"nama": nama, "stok": stok, "harga_beli": harga_beli, "harga": harga}).execute()
                            catat_log(nama, "Barang Baru", stok, f"Pendaftaran barang baru dengan stok awal {stok}")
                            st.success(f"Barang '{nama}' berhasil didaftarkan di Cloud!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Eror Sistem: {str(e)}")

            elif mode == "Update Stok Masuk/Keluar":
                df_pilihan = ambil_data()
                if df_pilihan.empty:
                    st.warning("Belum ada data barang di database.")
                else:
                    pilihan_barang = st.selectbox("Pilih Barang:", df_pilihan["nama"].tolist())
                    data_barang = df_pilihan[df_pilihan["nama"] == pilihan_barang].iloc

                    st.info(f"Stok: **{data_barang['stok']} Pcs** | Modal: **Rp {float(data_barang['harga_beli']):,.0f}** | Jual: **Rp {float(data_barang['harga']):,.0f}**")
                    jenis_opsi = st.selectbox("Jenis Mutasi:", ["Stok Masuk (+)", "Stok Keluar (-)"])
                    jumlah_mutasi = st.number_input("Jumlah Perubahan Stok", min_value=1, step=1)
                    keterangan = st.text_input("Keterangan / Catatan tambahan", placeholder="Contoh: Restock Supplier")
                    harga_jual_baru = st.number_input("Perbarui Harga Jual (Biarkan jika tetap)", value=float(data_barang["harga"]))

                    if st.button("🔄 Proses Perubahan", type="primary"):
                        stok_akhir = int(data_barang["stok"])
                        tipe_log = "Masuk" if jenis_opsi == "Stok Masuk (+)" else "Keluar"
                        stok_akhir = stok_akhir + jumlah_mutasi if jenis_opsi == "Stok Masuk (+)" else stok_akhir - jumlah_mutasi

                        if stok_akhir < 0:
                            st.error("Gagal! Stok tidak boleh kurang dari 0.")
                        else:
                            try:
                                supabase.table("barang").update({"stok": stok_akhir, "harga": harga_jual_baru}).eq("id", int(data_barang["id"])).execute()
                                catat_log(pilihan_barang, tipe_log, jumlah_mutasi, keterangan)
                                st.success("Stok berhasil diperbarui!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Eror Perubahan: {str(e)}")

            elif mode == "Hapus Barang":
                df_pilihan = ambil_data()
                if df_pilihan.empty:
                    st.warning("Belum ada data barang.")
                else:
                    pilihan_barang = st.selectbox("Pilih Barang yang akan dihapus:", df_pilihan["nama"].tolist())
                    if st.button("🗑️ Hapus Permanen", type="secondary"):
                        try:
                            supabase.table("barang").delete().eq("nama", pilihan_barang).execute()
                            catat_log(pilihan_barang, "Hapus", 0, "Barang dihapus permanen dari sistem")
                            st.success("Barang berhasil dihapus!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Eror Hapus: {str(e)}")
        with kolom_kanan:
            st.subheader("📋 Daftar Stok Gudang Real-time")
            cari_input = st.text_input("🔍 Cari Nama Barang...", placeholder="Ketik untuk memfilter...")
            df_stok = ambil_data(cari_input)

            if not df_stok.empty:
                df_tampil = df_stok.copy()
                df_tampil.columns = ["ID", "Nama Barang", "Stok", "Harga Modal (Rp)", "Harga Jual (Rp)"]

                stok_kritis = df_tampil[df_tampil["Stok"] <= 2]
                if not stok_kritis.empty:
                    st.error(f"🚨 **Peringatan:** Ada {len(stok_kritis)} barang dengan stok kritis (≤ 2 pcs)!", icon="⚠️")

                df_tampil["Harga Modal (Rp)"] = df_tampil["Harga Modal (Rp)"].apply(lambda x: f"Rp {x:,.0f}".replace(",", "."))
                df_tampil["Harga Jual (Rp)"] = df_tampil["Harga Jual (Rp)"].apply(lambda x: f"Rp {x:,.0f}".replace(",", "."))
                st.table(df_tampil)

                st.markdown("---")
                
                total_modal = (df_stok["stok"] * df_stok["harga_beli"]).sum()
                total_omzet = (df_stok["stok"] * df_stok["harga"]).sum()
                total_profit_bersih = total_omzet - total_modal

                txt_modal = f"Rp {total_modal:,.0f}".replace(",", ".")
                txt_omzet = f"Rp {total_omzet:,.0f}".replace(",", ".")
                txt_profit = f"Rp {total_profit_bersih:,.0f}".replace(",", ".")

                # Kartu Dashboard Finansial Responsif HTML & CSS (Anti-Terpotong di HP)
                html_kartu_keuangan = f"""
                <div style="display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 20px; font-family: sans-serif;">
                    <div style="flex: 1; min-width: 140px; background: linear-gradient(135deg, #2B6CB0, #4299E1); padding: 15px; border-radius: 12px; color: white;">
                        <div style="font-size: 11px; text-transform: uppercase; opacity: 0.9;">📦 Total Stok</div>
                        <div style="font-size: 18px; font-weight: bold; margin-top: 5px;">{df_stok['stok'].sum()} <span style="font-size: 12px; font-weight: normal;">Pcs</span></div>
                    </div>
                    <div style="flex: 1; min-width: 140px; background: linear-gradient(135deg, #4A5568, #718096); padding: 15px; border-radius: 12px; color: white;">
                        <div style="font-size: 11px; text-transform: uppercase; opacity: 0.9;">📉 Total Modal</div>
                        <div style="font-size: 16px; font-weight: bold; margin-top: 5px;">{txt_modal}</div>
                    </div>
                    <div style="flex: 1; min-width: 140px; background: linear-gradient(135deg, #D69E2E, #ECC94B); padding: 15px; border-radius: 12px; color: white;">
                        <div style="font-size: 11px; text-transform: uppercase; opacity: 0.9;">📈 Potensi Omzet</div>
                        <div style="font-size: 16px; font-weight: bold; margin-top: 5px;">{txt_omzet}</div>
                    </div>
                    <div style="flex: 1; min-width: 140px; background: linear-gradient(135deg, #2F855A, #48BB78); padding: 15px; border-radius: 12px; color: white;">
                        <div style="font-size: 11px; text-transform: uppercase; opacity: 0.9;">💰 Estimasi Profit</div>
                        <div style="font-size: 16px; font-weight: bold; margin-top: 5px;">{txt_profit}</div>
                    </div>
                </div>
                """
                st.markdown(html_kartu_keuangan, unsafe_allow_html=True)

                st.markdown("### 📊 Grafik Perbandingan Kuantitas Stok")
                df_grafik = df_stok[["nama", "stok"]].copy()
                df_grafik.columns = ["Nama Barang", "Jumlah Stok"]
                st.bar_chart(data=df_grafik, x="Nama Barang", y="Jumlah Stok", color="#2B6CB0")

                st.markdown("---")
                with st.container():
                    data_excel = konversi_ke_excel(df_tampil, "Daftar Stok")
                    st.download_button(label="🟢 Ekspor ke Excel (.xlsx)", data=data_excel, file_name="laporan_stok_barang.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else:
                st.info("Tidak ada data barang ditemukan.")

    elif menu == "📋 Riwayat / Log Transaksi":
        st.title("📋 Log & Riwayat Mutasi Stok")
        st.markdown("---")
        df_riwayat = ambil_riwayat()

        if not df_riwayat.empty:
            df_riwayat_tampil = df_riwayat.copy()
            df_riwayat_tampil.columns = ["Waktu & Tanggal", "Nama Barang", "Tipe Aktivitas", "Jumlah (Pcs)", "Keterangan / Catatan"]
            st.table(df_riwayat_tampil)
            st.markdown("---")
            
            st.subheader("🖨️ Cetak Nota Transaksi Keluar")
            df_keluar = df_riwayat_tampil[df_riwayat_tampil["Tipe Aktivitas"] == "Keluar"]
            
            if df_keluar.empty:
                st.info("Belum ada transaksi 'Stok Keluar' yang bisa dibuatkan nota.")
            else:
                opsi_nota = [f"{row['Waktu & Tanggal']} - {row['Nama Barang']} ({row['Jumlah (Pcs)']} Pcs)" for _, row in df_keluar.iterrows()]
                transaksi_terpilih = st.selectbox("Pilih Transaksi:", opsi_nota)
                
                indeks_pilihan = opsi_nota.index(transaksi_terpilih)
                data_nota = df_keluar.iloc[indeks_pilihan]
                
                html_nota = f"""
                <div style="font-family: 'Courier New', Courier, monospace; width: 280px; padding: 15px; border: 1px dashed #000; background-color: #fff; color: #000; margin: 0 auto;">
                    <div style="text-align: center; font-weight: bold; font-size: 14px;">NOTA PENGELUARAN STOK</div>
                    <div style="text-align: center; font-size: 11px; margin-bottom: 10px;">ENTERPRISE CLOUD SYSTEM</div>
                    <hr style="border-top: 1px dashed #000;">
                    <table style="width: 100%; font-size: 11px;">
                        <tr><td>Waktu</td><td>: {data_nota['Waktu & Tanggal']}</td></tr>
                        <tr><td>Barang</td><td>: {data_nota['Nama Barang']}</td></tr>
                        <tr><td>Jumlah</td><td>: {data_nota['Jumlah (Pcs)']} Pcs</td></tr>
                    </table>
                    <hr style="border-top: 1px dashed #000;">
                    <div style="font-size: 11px; word-wrap: break-word;"><strong>Catatan:</strong><br>{data_nota['Keterangan / Catatan']}</div>
                    <hr style="border-top: 1px dashed #000;">
                    <div style="text-align: center; font-size: 10px; margin-top: 10px;">Terima kasih.<br>Dokumen sah sistem cloud.</div>
                </div>
                """
                st.markdown("### 🔍 Pratinjau Nota:")
                st.html(html_nota)
                
                st.markdown("<br>", unsafe_allow_html=True)
                js_cetak = "<script>function cetakNota(){ window.print(); }</script><button onclick='cetakNota()' style='width: 100%; background-color: #ff4b4b; color: white; border: none; padding: 10px; border-radius: 8px; font-weight: bold; cursor: pointer;'>🖨️ Cetak / Simpan ke PDF</button>"
                st.components.v1.html(js_cetak, height=50)

            st.markdown("---")
            st.markdown("### 🧠 Rekomendasi Analisis AI Gudang (Real-time)")
            df_analisis_keluar = df_riwayat_tampil[df_riwayat_tampil["Tipe Aktivitas"] == "Keluar"]
            
            if df_analisis_keluar.empty:
                st.info("🤖 AI belum mendeteksi transaksi 'Stok Keluar' untuk dianalisis.")
            else:
                try:
                    df_laris = df_analisis_keluar.groupby("Nama Barang")["Jumlah (Pcs)"].sum().reset_index()
                    df_laris = df_laris.sort_values(by="Jumlah (Pcs)", ascending=False)
                    barang_paling_laris = df_laris.iloc["Nama Barang"]
                    jumlah_paling_laris = df_laris.iloc["Jumlah (Pcs)"]
                    
                    st.success(f"""
                    **🤖 Laporan Analisis AI Gudang:**
                    * 📈 **Produk Terlaris (Fast Moving):** Produk **'{barang_paling_laris}'** mencatat perputaran tertinggi dengan total **{jumlah_paling_laris} Pcs** keluar.
                    * 💡 **Rekomendasi Finansial AI:** Berdasarkan sisa margin profit bersih, disarankan menambah kuota belanja produk **'{barang_paling_laris}'** sebesar 20% bulan ini untuk memaksimalkan omzet toko Anda.
                    """)
                except Exception:
                    st.info("🤖 AI sedang menyusun analisis riwayat cloud...")

            st.markdown("---")
            data_excel_log = konversi_ke_excel(df_riwayat_tampil, "Log Riwayat")
            st.download_button(label="🟢 Unduh Seluruh Log Riwayat (.xlsx)", data=data_excel_log, file_name="riwayat_mutasi_stok.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Belum ada riwayat aktivitas.")
